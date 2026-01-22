import openpyxl
from openpyxl.styles import Font

def chuyen_doi_font(file_excel, sheet_name, font_moi):
    # Mở tệp Excel
    wb = openpyxl.load_workbook(file_excel)

    # Chọn sheet cần chuyển đổi font
    sheet = wb[page 1]

    # Duyệt qua từng dòng và cột
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            # Chuyển đổi font
            cell.font = Font(name=font_moi)

    # Lưu lại các thay đổi
    wb.save(file_excel)

# Sử dụng hàm
file_excel = "duong_dan/tên_tệp.xlsx"  # Điền đúng đường dẫn và tên tệp Excel của bạn
sheet_name = "Sheet1"  # Điền tên sheet bạn muốn chuyển đổi font
font_moi = "Arial"  # Điền tên font mới bạn muốn sử dụng

chuyen_doi_font(file_excel, sheet_name, font_moi)