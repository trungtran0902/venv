import time
import csv
import os
import re
import json
from datetime import datetime
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

# ======================
# 🧭 PARSE lat,lng
# ======================
def parse_latlng(text):
    match = re.match(r"\s*(-?\d+(\.\d+)?)\s*,\s*(-?\d+(\.\d+)?)\s*", str(text))
    if not match:
        return None, None
    return match.group(1), match.group(3)

# ======================
# 🧼 CLEAN CELL EXCEL
# ======================
def clean_cell(value):
    if value is None:
        return ""
    return str(value).strip()

# ======================
# 📞 NORMALIZE PHONE VN
# ======================
def normalize_phone_vn(phone):
    if not phone:
        return ""
    digits = re.sub(r"\D", "", phone)
    if digits.startswith("84"):
        return f"+{digits}"
    if digits.startswith("0"):
        return f"+84{digits[1:]}"
    return f"+{digits}"

# ======================
# 🧭 TÁCH lat,lng TỪ URL
# ======================
def extract_latlng_from_url(url):
    match = re.search(r"@(-?\d+\.\d+),(-?\d+\.\d+)", url)
    return (match.group(1), match.group(2)) if match else ("", "")

# ======================
# 🌀 SCROLL DANH SÁCH
# ======================
def scroll_results(page, max_round=40):
    feed = page.query_selector("div[role='feed']")
    if not feed:
        print("❌ Không tìm thấy danh sách kết quả")
        return False

    last_height = 0
    for _ in range(max_round):
        page.evaluate("(el) => el.scrollBy(0, el.scrollHeight)", feed)
        time.sleep(1.2)
        new_height = page.evaluate("(el) => el.scrollHeight", feed)
        if new_height == last_height:
            break
        last_height = new_height

    return True

# ======================
# 📍 ADDRESS
# ======================
def get_address(page):
    selectors = [
        "button[data-item-id='address']",
        "button[aria-label^='Địa chỉ']",
        "div[aria-label^='Địa chỉ']"
    ]
    for sel in selectors:
        loc = page.locator(sel)
        if loc.count() > 0:
            return loc.first.text_content().strip()
    return ""

# ======================
# 📞 PHONE
# ======================
def get_phone(page):
    try:
        btn = page.locator("button[data-item-id^='phone']")
        if btn.count() > 0:
            return btn.first.text_content().strip()
    except:
        pass
    return ""

# ======================
# 🌐 WEBSITE
# ======================
def get_website(page):
    try:
        link = page.locator("a[data-item-id='authority']")
        if link.count() > 0:
            return link.first.get_attribute("href")
    except:
        pass
    return ""

# ======================
# ⏰ OPENING HOURS TODAY
# ======================
def get_opening_hours_today(page):
    try:
        oh = page.locator("div[data-item-id='oh']")
        if oh.count() > 0:
            return oh.first.text_content().strip()
    except:
        pass
    return ""

# ======================
# ⏰ OPEN POPUP GIỜ (FIX)
# ======================
def open_opening_hours_popup(page):
    selectors = [
        "div[data-item-id='oh']",
        "button[data-item-id='oh']",
        "[aria-label*='Giờ']",
        "[aria-label*='Mở cửa']",
        "[aria-label*='Đóng cửa']",
        "[aria-label*='Open']",
        "[aria-label*='Closed']",
    ]

    for sel in selectors:
        loc = page.locator(sel)
        if loc.count() > 0:
            try:
                loc.first.click(force=True, timeout=1500)
                page.wait_for_selector("div[role='dialog']", timeout=3000)
                return True
            except:
                pass
    return False

# ======================
# ⏰ FULL WEEK HOURS (FIX)
# ======================
def get_opening_hours_full_week(page):
    hours = {}

    dialog = page.locator("div[role='dialog']").last
    if dialog.count() == 0:
        return hours

    # Case A: TABLE
    try:
        rows = dialog.locator("table tr")
        for i in range(rows.count()):
            cols = rows.nth(i).locator("td")
            if cols.count() >= 2:
                day = (cols.nth(0).text_content() or "").strip()
                t = (cols.nth(1).text_content() or "").strip()
                if day and t:
                    hours[day] = t
        if hours:
            return hours
    except:
        pass

    # Case B: DIV rows
    try:
        rows = dialog.locator("[role='row']")
        for i in range(rows.count()):
            cells = rows.nth(i).locator("[role='cell']")
            if cells.count() >= 2:
                day = (cells.nth(0).text_content() or "").strip()
                t = (cells.nth(1).text_content() or "").strip()
                if day and t:
                    hours[day] = t
        if hours:
            return hours
    except:
        pass

    # Case C: TEXT fallback
    try:
        text = (dialog.text_content() or "").strip()
        for line in [x.strip() for x in text.splitlines() if x.strip()]:
            if line.startswith("Thứ") or line.startswith("Chủ"):
                m = re.match(r"^(Thứ\s+\w+|Chủ\s+Nhật)\s+(.+)$", line)
                if m:
                    hours[m.group(1)] = m.group(2)
    except:
        pass

    return hours

# ======================
# 📍 PARSE SINGLE POI
# ======================
def parse_current_poi(page, center_lat, center_lng, keyword):
    try:
        name = page.locator("h1.DUwDvf").first.text_content()
    except:
        return None

    address = get_address(page)
    phone = normalize_phone_vn(get_phone(page))
    website = get_website(page)
    opening_today = get_opening_hours_today(page)

    opening_full = {}
    if open_opening_hours_popup(page):
        opening_full = get_opening_hours_full_week(page)
        page.keyboard.press("Escape")
        time.sleep(0.3)

    lat, lng = extract_latlng_from_url(page.url)

    return {
        "center_lat": center_lat,
        "center_lng": center_lng,
        "keyword": keyword,
        "name": name,
        "address": address,
        "phone": phone,
        "website": website,
        "opening_hours_today": opening_today,
        "opening_hours_full": json.dumps(opening_full, ensure_ascii=False),
        "lat": lat,
        "lng": lng,
        "url": page.url
    }

# ======================
# 🚀 CRAWL GOOGLE MAPS
# ======================
def crawl_google_maps(center_lat, center_lng, keyword):
    results = []

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=False,
            slow_mo=100,
            args=["--disable-blink-features=AutomationControlled"]
        )

        context = browser.new_context(
            locale="vi-VN",
            viewport={"width": 1280, "height": 800},
            extra_http_headers={"Accept-Language": "vi"}
        )

        page = context.new_page()

        print(f"📍 Di chuyển tới: {center_lat},{center_lng}")
        page.goto(
            f"https://www.google.com/maps/@{center_lat},{center_lng},14z?hl=vi",
            timeout=60000
        )
        time.sleep(5)

        print(f"🔎 Tìm kiếm: {keyword}")
        search_box = page.wait_for_selector("input[role='combobox']", timeout=10000)
        search_box.fill(keyword)
        search_box.press("Enter")
        time.sleep(6)

        print("🌀 Cuộn danh sách...")
        has_feed = scroll_results(page)

        links = []
        if has_feed:
            links = page.eval_on_selector_all(
                "a[href*='/maps/place/']",
                "els => [...new Set(els.map(el => el.href))]"
            )

        if not links:
            print("ℹ️ Single POI detected")
            poi = parse_current_poi(page, center_lat, center_lng, keyword)
            if poi:
                results.append(poi)
            browser.close()
            return results

        print(f"📌 Lấy được {len(links)} địa điểm")

        for idx, link in enumerate(links):
            try:
                page.goto(link, timeout=60000)
                time.sleep(3)

                poi = parse_current_poi(page, center_lat, center_lng, keyword)
                if poi:
                    results.append(poi)
                    print(f"✔ {idx+1:03d}: {poi['name']}")

            except Exception as e:
                print(f"❌ Lỗi {idx+1}: {e}")

        browser.close()
    return results

# ======================
# 💾 SAVE CSV
# ======================
def save_all_to_csv(output_file, all_data):
    if not all_data:
        print("⚠️ Không có dữ liệu để ghi")
        return

    with open(output_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "center_lat", "center_lng",
                "keyword",
                "name", "address",
                "phone", "website",
                "opening_hours_today",
                "opening_hours_full",
                "lat", "lng",
                "url"
            ]
        )
        writer.writeheader()
        writer.writerows(all_data)

# ======================
# ▶️ MAIN
# ======================
if __name__ == "__main__":

    folder = input("📂 Nhập đường dẫn folder Excel: ").strip()
    file = input("📄 Nhập tên file Excel: ").strip()
    path = os.path.join(folder, file)

    wb = load_workbook(path)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    print("\n📊 Các cột:")
    for i, h in enumerate(headers, 1):
        print(f"{i}. {h}")

    loc_idx = int(input("\n📍 STT cột location: ")) - 1
    key_idx = int(input("🔎 STT cột keyword: ")) - 1

    jobs = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        location = clean_cell(row[loc_idx])
        keyword = clean_cell(row[key_idx])
        if not location or not keyword:
            print(f"⚠️ Bỏ qua dòng {row_num}")
            continue
        jobs.append((location, keyword))

    print(f"\n🚀 Tổng job: {len(jobs)}")

    all_results = []
    for i, (loc, key) in enumerate(jobs, 1):
        print(f"\n▶️ Job {i}/{len(jobs)}: {key}")
        lat, lng = parse_latlng(loc)
        if not lat:
            print("❌ Sai định dạng lat,lng")
            continue
        all_results.extend(crawl_google_maps(lat, lng, key))
        time.sleep(8)

    out = f"googlemaps_full_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    save_all_to_csv(out, all_results)

    print(f"\n✅ Hoàn tất {len(all_results)} records")
    print(f"📁 File: {out}")
