import time
import csv
import os
import re
import json
import math
from datetime import datetime
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

# ======================
# ⚙️ CONFIG
# ======================
MAX_RADIUS_METERS = 50
MAX_RADIUS_KM = MAX_RADIUS_METERS / 1000

# ======================
# 🧭 PARSE lat,lng
# ======================
def parse_latlng(text):
    match = re.match(r"\s*(-?\d+(\.\d+)?)\s*,\s*(-?\d+(\.\d+)?)\s*", str(text))
    if not match:
        return None, None
    return match.group(1), match.group(3)

# ======================
# 🧼 CLEAN CELL EXCEL
# ======================
def clean_cell(value):
    if value is None:
        return ""
    return str(value).strip()

# ======================
# 📞 NORMALIZE PHONE VN
# ======================
def normalize_phone_vn(phone):
    if not phone:
        return ""
    digits = re.sub(r"\D", "", phone)
    if digits.startswith("84"):
        return f"+{digits}"
    if digits.startswith("0"):
        return f"+84{digits[1:]}"
    return f"+{digits}"

# ======================
# 🧭 TÁCH lat,lng TỪ URL
# ======================
def extract_latlng_from_url(url):
    match = re.search(r"@(-?\d+\.\d+),(-?\d+\.\d+)", url)
    return (match.group(1), match.group(2)) if match else ("", "")

# ======================
# 📏 HAVERSINE DISTANCE
# ======================
def haversine_distance(lat1, lng1, lat2, lng2):
    R = 6371
    lat1, lng1, lat2, lng2 = map(
        math.radians, [lat1, lng1, lat2, lng2]
    )
    dlat = lat2 - lat1
    dlng = lng2 - lng1
    a = math.sin(dlat/2)**2 + \
        math.cos(lat1)*math.cos(lat2)*math.sin(dlng/2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    return R * c

# ======================
# 🌀 SCROLL DANH SÁCH
# ======================
def scroll_results(page, max_round=40):
    feed = page.query_selector("div[role='feed']")
    if not feed:
        print("❌ Không tìm thấy danh sách kết quả")
        return False
    last_height = 0
    for _ in range(max_round):
        page.evaluate("(el) => el.scrollBy(0, el.scrollHeight)", feed)
        time.sleep(1.2)
        new_height = page.evaluate("(el) => el.scrollHeight", feed)
        if new_height == last_height:
            break
        last_height = new_height
    return True

# ======================
# 📍 ADDRESS
# ======================
def get_address(page):
    selectors = [
        "button[data-item-id='address']",
        "button[aria-label^='Địa chỉ']",
        "div[aria-label^='Địa chỉ']"
    ]
    for sel in selectors:
        loc = page.locator(sel)
        if loc.count() > 0:
            return loc.first.text_content().strip()
    return ""

# ======================
# 📞 PHONE
# ======================
def get_phone(page):
    btn = page.locator("button[data-item-id^='phone']")
    return btn.first.text_content().strip() if btn.count() > 0 else ""

# ======================
# 🌐 WEBSITE
# ======================
def get_website(page):
    link = page.locator("a[data-item-id='authority']")
    return link.first.get_attribute("href") if link.count() > 0 else ""

# ======================
# ⏰ OPENING HOURS TODAY
# ======================
def get_opening_hours_today(page):
    oh = page.locator("div[data-item-id='oh']")
    return oh.first.text_content().strip() if oh.count() > 0 else ""

# ======================
# ⏰ OPEN POPUP GIỜ
# ======================
def open_opening_hours_popup(page):
    selectors = [
        "div[data-item-id='oh']",
        "button[data-item-id='oh']",
        "[aria-label*='Giờ']",
        "[aria-label*='Mở cửa']",
        "[aria-label*='Đóng cửa']",
    ]
    for sel in selectors:
        loc = page.locator(sel)
        if loc.count() > 0:
            try:
                loc.first.click(force=True)
                page.wait_for_selector("div[role='dialog']", timeout=3000)
                return True
            except:
                pass
    return False

# ======================
# ⏰ FULL WEEK HOURS (WAIT)
# ======================
def get_opening_hours_full_week(page):
    hours = {}
    dialog = page.locator("div[role='dialog']").last
    if dialog.count() == 0:
        return hours

    try:
        page.wait_for_selector(
            "div[role='dialog'] table tr, div[role='dialog'] [role='row']",
            timeout=3000
        )
    except:
        pass

    rows = dialog.locator("table tr")
    if rows.count() > 0:
        for i in range(rows.count()):
            cols = rows.nth(i).locator("td")
            if cols.count() >= 2:
                d = cols.nth(0).text_content().strip()
                t = cols.nth(1).text_content().strip()
                if d and t:
                    hours[d] = t
        return hours

    rows = dialog.locator("[role='row']")
    for i in range(rows.count()):
        cells = rows.nth(i).locator("[role='cell']")
        if cells.count() >= 2:
            d = cells.nth(0).text_content().strip()
            t = cells.nth(1).text_content().strip()
            if d and t:
                hours[d] = t
    return hours

# ======================
# 📍 PARSE POI
# ======================
def parse_current_poi(page, center_lat, center_lng, keyword):
    try:
        name = page.locator("h1.DUwDvf").first.text_content()
    except:
        return None

    lat, lng = extract_latlng_from_url(page.url)
    if not lat or not lng:
        return None

    distance_km = haversine_distance(
        float(center_lat), float(center_lng),
        float(lat), float(lng)
    )

    opening_full = {}
    if open_opening_hours_popup(page):
        time.sleep(0.8)
        opening_full = get_opening_hours_full_week(page)
        page.keyboard.press("Escape")
        time.sleep(0.3)

    return {
        "center_lat": center_lat,
        "center_lng": center_lng,
        "keyword": keyword,
        "name": name,
        "address": get_address(page),
        "phone": normalize_phone_vn(get_phone(page)),
        "website": get_website(page),
        "opening_hours_today": get_opening_hours_today(page),
        "opening_hours_full": json.dumps(opening_full, ensure_ascii=False),
        "lat": lat,
        "lng": lng,
        "distance_km": round(distance_km, 4),
        "url": page.url
    }

# ======================
# 🚀 CRAWL GOOGLE MAPS
# ======================
def crawl_google_maps(center_lat, center_lng, keyword):
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, slow_mo=100)
        page = browser.new_page()

        page.goto(
            f"https://www.google.com/maps/@{center_lat},{center_lng},15z?hl=vi",
            timeout=60000
        )
        time.sleep(4)

        search = page.wait_for_selector("input[role='combobox']")
        search.fill(keyword)
        search.press("Enter")
        time.sleep(6)

        has_feed = scroll_results(page)

        links = []
        if has_feed:
            links = page.eval_on_selector_all(
                "a[href*='/maps/place/']",
                "els => [...new Set(els.map(el => el.href))]"
            )

        closest = None
        min_dist = float("inf")

        def consider(poi):
            nonlocal closest, min_dist
            d = poi["distance_km"]
            if d <= MAX_RADIUS_KM and d < min_dist:
                min_dist = d
                closest = poi

        if not links:
            poi = parse_current_poi(page, center_lat, center_lng, keyword)
            if poi:
                consider(poi)
            browser.close()
            return [closest] if closest else []

        for link in links:
            page.goto(link)
            time.sleep(3)
            poi = parse_current_poi(page, center_lat, center_lng, keyword)
            if poi:
                consider(poi)

        browser.close()
        return [closest] if closest else []

# ======================
# 💾 SAVE CSV
# ======================
def save_all_to_csv(output_file, data):
    if not data:
        print("⚠️ Không có dữ liệu")
        return
    with open(output_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=list(data[0].keys())
        )
        writer.writeheader()
        writer.writerows(data)

# ======================
# ▶️ MAIN
# ======================
if __name__ == "__main__":

    folder = input("📂 Folder Excel: ").strip()
    file = input("📄 Tên file Excel: ").strip()
    path = os.path.join(folder, file)

    wb = load_workbook(path)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    for i, h in enumerate(headers, 1):
        print(f"{i}. {h}")

    loc_idx = int(input("📍 STT cột location: ")) - 1
    key_idx = int(input("🔎 STT cột keyword: ")) - 1

    jobs = []
    for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        loc = clean_cell(row[loc_idx])
        key = clean_cell(row[key_idx])
        if loc and key:
            jobs.append((loc, key))
        else:
            print(f"⚠️ Bỏ dòng {r}")

    all_results = []
    for loc, key in jobs:
        lat, lng = parse_latlng(loc)
        if not lat:
            continue
        res = crawl_google_maps(lat, lng, key)
        all_results.extend(res)
        time.sleep(6)

    out = f"googlemaps_radius50m_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    save_all_to_csv(out, all_results)

    print(f"\n✅ Hoàn tất: {len(all_results)} record")
    print(f"📁 File: {out}")
