import time, os, re, json, math
from datetime import datetime
from openpyxl import load_workbook, Workbook
from playwright.sync_api import sync_playwright
import tkinter as tk
from tkinter import filedialog

# ======================
# UTILS
# ======================
def clean_cell(v):
    return "" if v is None else str(v).strip()

def normalize_phone_vn(phone):
    if not phone:
        return ""
    d = re.sub(r"\D", "", phone)
    if not d:
        return ""
    if d.startswith("84"):
        return f"+{d}"
    if d.startswith("0"):
        return f"+84{d[1:]}"
    return f"+{d}"

def extract_latlng_from_url(url):
    m = re.search(r"!3d(-?\d+(?:\.\d+)?)!4d(-?\d+(?:\.\d+)?)", url)
    if m:
        return m.group(1), m.group(2)

    m = re.search(r"@(-?\d+(?:\.\d+)?),(-?\d+(?:\.\d+)?)", url)
    if m:
        return m.group(1), m.group(2)

    return "", ""

def distance_km(lat1, lon1, lat2, lon2):
    R = 6371
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (
        math.sin(dlat / 2) ** 2
        + math.cos(math.radians(lat1))
        * math.cos(math.radians(lat2))
        * math.sin(dlon / 2) ** 2
    )
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c

# ======================
# ENSURE SEARCH BOX
# ======================
SEARCH_SELECTORS = [
    "input[role='combobox']",
    "input[aria-label*='Tìm kiếm']",
    "input[aria-label*='Search']",
]

def ensure_search_box(page, timeout=15000):
    for s in SEARCH_SELECTORS:
        try:
            loc = page.wait_for_selector(s, timeout=3000, state="visible")
            if loc:
                return loc
        except:
            pass

    print("⚠️ Không thấy search box → reload Google Maps")
    page.goto("https://www.google.com/maps?hl=vi", timeout=60000)
    time.sleep(3)

    return page.wait_for_selector(
        "input[role='combobox']",
        timeout=timeout,
        state="visible"
    )

# ======================
# SCROLL
# ======================
def scroll_results(page, max_rounds=30):
    feed = page.locator("div[role='feed']")
    if feed.count() == 0:
        return False

    last = 0
    for _ in range(max_rounds):
        try:
            feed.first.evaluate("(el)=>el.scrollBy(0, el.scrollHeight)")
            time.sleep(0.4)
            h = feed.first.evaluate("(el)=>el.scrollHeight")
            if h == last:
                break
            last = h
        except:
            break
    return True

def scroll_detail_panel(page, max_rounds=8):
    panel = page.locator("div[role='main']")
    if panel.count() == 0:
        return
    last = 0
    for _ in range(max_rounds):
        try:
            panel.first.evaluate("(el)=>el.scrollBy(0, el.scrollHeight)")
            time.sleep(0.2)
            h = panel.first.evaluate("(el)=>el.scrollHeight")
            if h == last:
                break
            last = h
        except:
            break

# ======================
# BASIC FIELDS
# ======================
def get_address(page):
    for s in [
        "button[data-item-id='address']",
        "button[aria-label^='Địa chỉ']",
        "div[aria-label^='Địa chỉ']",
    ]:
        loc = page.locator(s)
        if loc.count() > 0:
            return (loc.first.text_content() or "").strip()
    return ""

def get_phone(page):
    loc = page.locator("button[data-item-id^='phone']")
    return (loc.first.text_content() or "").strip() if loc.count() > 0 else ""

def get_website(page):
    loc = page.locator("a[data-item-id='authority']")
    return loc.first.get_attribute("href") if loc.count() > 0 else ""

def get_located_in(page):
    try:
        loc = page.locator("text=/Nằm ở:/")
        if loc.count() > 0:
            txt = loc.first.text_content() or ""
            return txt.replace("Nằm ở:", "").strip()
    except:
        pass
    return ""

# ======================
# OPEN STATUS
# ======================
def get_open_status(page):
    try:
        panel = page.locator("div[role='main']")
        if panel.count() == 0:
            return ""

        txt = (panel.first.text_content() or "").lower()

        if "đóng cửa vĩnh viễn" in txt:
            return "đóng cửa vĩnh viễn"
        if "sắp mở cửa" in txt:
            return "sắp mở cửa"
        if "đang mở cửa" in txt:
            return "đang mở cửa"
        if "đóng cửa" in txt:
            return "đóng cửa"

        return ""
    except:
        return ""

# ======================
# FILTER BASIC
# ======================
def get_basic_poi_for_filter(page):
    try:
        name = (page.locator("h1").first.text_content() or "").strip()
        lat, lng = extract_latlng_from_url(page.url)
        return name, lat, lng
    except:
        return None, "", ""

# ======================
# PARSE FULL POI
# ======================
def parse_current_poi(page, keyword):
    try:
        name = (page.locator("h1").first.text_content() or "").strip()
        if not name:
            return None
    except:
        return None

    scroll_detail_panel(page)

    address = get_address(page)
    located_in = get_located_in(page)
    phone = normalize_phone_vn(get_phone(page))
    website = get_website(page)
    open_status = get_open_status(page)
    lat, lng = extract_latlng_from_url(page.url)

    return {
        "keyword": keyword,
        "name": name,
        "address": address,
        "located_in": located_in,
        "phone": phone,
        "website": website,
        "open_status": open_status,
        "lat": lat,
        "lng": lng,
        "url": page.url,
    }

# ======================
# LOAD DONE KEYWORDS
# ======================
def load_done_keywords(autosave_path):
    done = set()
    if not os.path.exists(autosave_path):
        return done

    try:
        wb = load_workbook(autosave_path)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        if "keyword" not in headers:
            return done

        k_idx = headers.index("keyword")
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r[k_idx]:
                done.add(str(r[k_idx]).strip().lower())
    except:
        pass

    return done

# ======================
# GET LINKS
# ======================
def get_place_links_from_list(page):
    try:
        return page.eval_on_selector_all(
            "a[href*='/maps/place/']",
            "els=>[...new Set(els.map(e=>e.href))]"
        )
    except:
        return []

# ======================
# FOCUS LOCATION
# ======================
def focus_location_with_radius(page, loc_text, radius_km=1.0):
    try:
        lat, lng = [x.strip() for x in loc_text.split(",")]
        zoom = 16 if radius_km <= 1 else 15 if radius_km <= 2 else 14
        page.goto(f"https://www.google.com/maps/@{lat},{lng},{zoom}z", timeout=60000)
        time.sleep(2)
        return float(lat), float(lng)
    except:
        return None, None

# ======================
# XLSX
# ======================
FIELDS = [
    "keyword", "name", "address", "located_in",
    "phone", "website", "open_status",
    "lat", "lng", "url"
]

def save_xlsx(path, data):
    wb = Workbook()
    ws = wb.active
    ws.append(FIELDS)
    for row in data:
        ws.append([row.get(k, "") for k in FIELDS])
    wb.save(path)

# ======================
# CRAWL KEYWORD
# ======================
def crawl_google_maps_keyword(page, keyword, center_lat=None, center_lng=None, radius_km=1.0):
    results = []

    sb = ensure_search_box(page)
    sb.click()
    sb.fill(keyword)
    sb.press("Enter")
    time.sleep(3)

    has_feed = scroll_results(page)
    links = get_place_links_from_list(page) if has_feed else [page.url]

    for i, l in enumerate(links, 1):
        try:
            page.goto(l, timeout=60000)
            time.sleep(1)

            name, lat, lng = get_basic_poi_for_filter(page)
            if not name:
                continue

            if center_lat and center_lng and lat and lng:
                d = distance_km(center_lat, center_lng, float(lat), float(lng))
                if d > radius_km:
                    print(f"   ⛔ Ngoài {radius_km}km:", name, round(d * 1000), "m")
                    continue

            poi = parse_current_poi(page, keyword)
            if poi:
                results.append(poi)
                print(f"✔ {keyword} | {i}: {poi['name']} → DỪNG keyword")
                break

        except Exception as e:
            print("❌", keyword, e)

    return results

# ======================
# MAIN
# ======================
if __name__ == "__main__":
    script_dir = os.path.dirname(os.path.abspath(__file__))

    root = tk.Tk()
    root.withdraw()
    path = filedialog.askopenfilename(
        title="Chọn file Excel",
        filetypes=[("Excel", "*.xlsx *.xls")]
    )
    if not path:
        print("❌ Chưa chọn file")
        exit()

    wb = load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    for i, h in enumerate(headers, 1):
        print(i, h)

    key_idx = int(input("🔎 Cột keyword: ")) - 1
    loc_idx = int(input("📍 Cột location (lat,lng): ")) - 1

    rows_data = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        kw = clean_cell(r[key_idx]) if key_idx < len(r) else ""
        loc = clean_cell(r[loc_idx]) if loc_idx < len(r) else ""
        if kw:
            rows_data.append((kw, loc))

    radius = 1.0
    autosave_path = os.path.join(script_dir, "autosave_temp.xlsx")
    done_keywords = load_done_keywords(autosave_path)

    all_results = []

    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            user_data_dir="google_profile",
            executable_path=r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            headless=False,
            locale="vi-VN",
            viewport={"width": 1280, "height": 800},
        )
        page = context.new_page()
        page.goto("https://www.google.com/maps?hl=vi", timeout=60000)
        time.sleep(3)

        for i, (k, loc) in enumerate(rows_data, 1):
            k_norm = k.strip().lower()
            if k_norm in done_keywords:
                print(f"⏭️ SKIP: {k}")
                continue

            print(f"\n▶️ {i}/{len(rows_data)}: {k}")

            page.goto("https://www.google.com/maps?hl=vi", timeout=60000)
            time.sleep(2)

            center_lat = center_lng = None
            if loc:
                center_lat, center_lng = focus_location_with_radius(page, loc, radius)

            results = crawl_google_maps_keyword(page, k, center_lat, center_lng, radius)
            all_results.extend(results)

            if all_results:
                save_xlsx(autosave_path, all_results)
                print("💾 Autosave")

        context.close()

    out = os.path.join(
        script_dir,
        f"googlemaps_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    save_xlsx(out, all_results)
    print("✅ DONE:", out)
