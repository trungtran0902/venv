import requests
import time
import random
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook, load_workbook
import os


# ======================
# COOKIE ƒêƒÇNG NH·∫¨P
# ======================
COOKIE = "..."   # üî¥ GI·ªÆ NGUY√äN COOKIE B·∫†N ƒê√É C√ì


# ======================
# CH·ªåN FILE TXT
# ======================
def choose_txt_file():
    root = tk.Tk()
    root.withdraw()

    file_path = filedialog.askopenfilename(
        title="Ch·ªçn file TXT ch·ª©a link API Foody",
        filetypes=[("Text files", "*.txt")]
    )
    return file_path


# ======================
# KH·ªûI T·∫†O / M·ªû FILE EXCEL
# ======================
def init_excel(file_name="foody_data.xlsx"):
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Foody Map Data"
        ws.append([
            "STT",
            "T√™n qu√°n",
            "ƒê·ªãa ch·ªâ",
            "ƒêi·ªán tho·∫°i",
            "Rating",
            "Latitude",
            "Longitude",
            "Distance",
            "IsDelivery",
            "IsOpening",
            "Google Maps"
        ])
        wb.save(file_name)

    return wb, ws


# ======================
# MAIN
# ======================
def main():
    # B1: ch·ªçn file TXT
    txt_file = choose_txt_file()
    if not txt_file:
        print("‚ùå Ch∆∞a ch·ªçn file TXT")
        return

    print("üìÇ File ƒë√£ ch·ªçn:", txt_file)

    # ƒë·ªçc danh s√°ch API
    with open(txt_file, "r", encoding="utf-8") as f:
        api_urls = [line.strip() for line in f if line.strip()]

    print(f"üîó T·ªïng link API: {len(api_urls)}\n")

    # session request
    s = requests.Session()
    s.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
        "Accept": "application/json, text/plain, */*",
        "Referer": "https://www.foody.vn/",
        "X-Requested-With": "XMLHttpRequest",
        "Cookie": COOKIE
    })

    # Excel
    excel_file = "foody_data_hanoi.xlsx"
    wb, ws = init_excel(excel_file)

    # s·ªë d√≤ng ƒë√£ c√≥
    total_place = ws.max_row - 1

    # B2: g·ªçi t·ª´ng API
    for idx, api_url in enumerate(api_urls, start=1):
        print(f"\nüåê [{idx}/{len(api_urls)}] GET: {api_url}")

        try:
            r = s.get(api_url, timeout=15)
            print("Status:", r.status_code)

            if r.status_code != 200:
                print("‚ùå L·ªói request")
                continue

            js = r.json()
            items = js.get("Items", [])

            print(f"S·ªë qu√°n trong batch: {len(items)}")

            for place in items:
                total_place += 1

                lat = place.get("Latitude")
                lng = place.get("Longitude")
                distance = place.get("Distance")
                is_delivery = place.get("IsDelivery")
                is_opening = place.get("IsOpening")

                maps_url = ""
                if lat and lng:
                    maps_url = f"https://www.google.com/maps?q={lat},{lng}"

                ws.append([
                    total_place,
                    place.get("Name"),
                    place.get("Address"),
                    place.get("Phone", ""),
                    place.get("AvgRating"),
                    lat,
                    lng,
                    distance,
                    is_delivery,
                    is_opening,
                    maps_url
                ])

            # üíæ AUTO SAVE sau m·ªói API
            wb.save(excel_file)
            print("üíæ ƒê√£ l∆∞u Excel")

            # ‚è≥ Delay ng·∫´u nhi√™n 2‚Äì3s
            sleep_time = random.uniform(2, 3)
            print(f"‚è≥ Ngh·ªâ {sleep_time:.2f}s ƒë·ªÉ tr√°nh call API li√™n t·ª•c...")
            time.sleep(sleep_time)

        except Exception as e:
            print("‚ùå L·ªói:", e)

    print(f"\nüéâ DONE ‚Äì T·ªïng s·ªë qu√°n l·∫•y ƒë∆∞·ª£c: {total_place}")
    print(f"üìÅ File Excel: {excel_file}")


if __name__ == "__main__":
    main()
