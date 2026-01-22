import openpyxl
import requests
import pandas as pd
from datetime import datetime

# Cấu hình
excel_path = "KÊ HOẠCH CẬP NHẬP SÁP NHẬP TỈNH THÀNH.xlsx"
api_template = "https://api-private.map4d.vn/osm/administrator-osm/{id}?isPolygon=false&Key=nghiant"

# Mở file Excel
wb = openpyxl.load_workbook(excel_path)
results = []

# Duyệt qua từng sheet
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))

    if not rows or len(rows[0]) < 3:
        continue

    header = rows[0]
    try:
        id_col_index = header.index("Relation Id mới")
    except ValueError:
        continue

    for i, row in enumerate(rows[1:], start=2):
        relation_id_raw = row[id_col_index]
        if not relation_id_raw:
            continue

        relation_id = str(relation_id_raw).strip()
        if not relation_id.isdigit():
            continue

        url = api_template.format(id=relation_id)
        try:
            response = requests.get(url, timeout=10)
            if response.status_code == 200:
                data = response.json()
                name = data.get("name", {}).get("vi", "")
                status = "✅ OK"
            else:
                name = ""
                status = f"❌ Error {response.status_code}"
        except Exception as e:
            name = ""
            status = f"⚠️ Exception: {e}"

        results.append({
            "Sheet": sheet_name,
            "Relation Id mới": relation_id,
            "API URL": url,
            "Phường/Xã mới": name,
            "Status": status
        })

# Xuất ra Excel
output_filename = f"ket_qua_api_map4d_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
df = pd.DataFrame(results)
df.to_excel(output_filename, index=False)
print(f"✅ Đã xuất file: {output_filename}")
