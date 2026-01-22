import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ==============================
# B1 + B2: Nhập đường dẫn + tên file
# ==============================
folder = input("Nhập đường dẫn chứa file Excel: ").strip()
filename = input("Nhập tên file Excel (vd: data.xlsx): ").strip()
filepath = f"{folder}/{filename}"

# ==============================
# B3: Nhập tên sheet
# ==============================
sheetname = input("Nhập tên sheet cần xử lý: ").strip()

# ==============================
# Đọc sheet và lấy danh sách cột
# ==============================
df = pd.read_excel(filepath, sheet_name=sheetname)

print("\nDanh sách các cột trong sheet:")
for i, col in enumerate(df.columns):
    print(f"{i+1}. {col}")

cols_input = input("\nNhập số thứ tự các cột dùng để kiểm tra trùng (vd: 1,3,5): ").strip()
selected_indexes = [int(x) - 1 for x in cols_input.split(",")]
duplicate_cols = [df.columns[i] for i in selected_indexes]

print("\n✔ Cột dùng để check trùng:", duplicate_cols)

# ==============================
# B4: Xử lý trùng – Loại bỏ mọi loại NULL
# ==============================

def is_valid(value):
    """Trả về True nếu value KHÔNG NULL theo mọi dạng."""
    if pd.isna(value):   # NaN, None
        return False

    v = str(value).strip().lower()

    # các dạng NULL xảy ra nhiều trong file thực tế
    invalid_set = {"", "null", "none", "nan", "n/a", "na"}

    if v in invalid_set:
        return False

    return True


# Tạo mask: chỉ dòng có FULL dữ liệu thật sự
mask_full = df[duplicate_cols].apply(lambda row: all(is_valid(v) for v in row), axis=1)

# Tạo cột is_duplicate
df["is_duplicate"] = False

# Chỉ dòng đủ dữ liệu mới được check duplicate
dup_mask = df[mask_full].duplicated(subset=duplicate_cols, keep=False)
df.loc[mask_full, "is_duplicate"] = dup_mask

# Sắp xếp: dòng trùng lên đầu
df_sorted = df.sort_values("is_duplicate", ascending=False)

# ==============================
# B5: Ghi ra file mới với toàn bộ sheet
# ==============================
output_file = f"{folder}/output.xlsx"

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    orig = pd.ExcelFile(filepath)
    for sh in orig.sheet_names:
        if sh == sheetname:
            df_sorted.to_excel(writer, sheet_name=sh, index=False)
        else:
            tmp = pd.read_excel(filepath, sheet_name=sh)
            tmp.to_excel(writer, sheet_name=sh, index=False)

# ==============================
# B6: Highlight các dòng trùng
# ==============================
wb = load_workbook(output_file)
ws = wb[sheetname]

yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

for row in range(2, ws.max_row + 1):
    if df_sorted.iloc[row - 2]["is_duplicate"]:
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = yellow

# Xóa cột is_duplicate
dup_col_index = df_sorted.columns.get_loc("is_duplicate") + 1
ws.delete_cols(dup_col_index)

wb.save(output_file)

print("\n🎉 DONE BRO!")
print("✔ Đã BỎ QUA hết các trường NULL, 'Null', 'nan', 'none', ô trống")
print("✔ Chỉ record đầy đủ dữ liệu mới check trùng")
print("✔ Dòng trùng đưa lên đầu + highlight")
print("📌 File output:", output_file)
