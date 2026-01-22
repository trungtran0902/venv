import time
import os
import re
import json
import math
from datetime import datetime
from openpyxl import load_workbook, Workbook
from playwright.sync_api import sync_playwright
import tkinter as tk
from tkinter import filedialog

# ======================
# UTILS
# ======================

def clean_cell(v):
    return "" if v is None else str(v).strip()

def normalize_phone_vn(phone):
    if not phone:
        return ""
    d = re.sub(r"\D", "", phone)
    if not d:
        return ""
    if d.startswith("84"):
        return f"+{d}"
    if d.startswith("0"):
        return f"+84{d[1:]}"
    return f"+{d}"

def extract_latlng_from_url(url):
    # Ưu tiên lấy từ !3d và !4d (tọa độ POI thật)
    m = re.search(r"!3d(-?\d+(?:\.\d+)?)!4d(-?\d+(?:\.\d+)?)", url)
    if m:
        return m.group(1), m.group(2)

    # Fallback: lấy từ @lat,lng
    m = re.search(r"@(-?\d+(?:\.\d+)?),(-?\d+(?:\.\d+)?)", url)
    if m:
        return m.group(1), m.group(2)

    return "", ""

def distance_km(lat1, lon1, lat2, lon2):
    R = 6371
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (
        math.sin(dlat/2)**2 +
        math.cos(math.radians(lat1)) *
        math.cos(math.radians(lat2)) *
        math.sin(dlon/2)**2
    )
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    return R * c

# ======================
# SCROLL
# ======================

def scroll_results(page, max_rounds=30):
    feed = page.locator("div[role='feed']")
    if feed.count() == 0:
        return False
    last = 0
    for _ in range(max_rounds):
        try:
            feed.first.evaluate("(el)=>el.scrollBy(0, el.scrollHeight)")
            time.sleep(0.4)
            h = feed.first.evaluate("(el)=>el.scrollHeight")
            if h == last:
                break
            last = h
        except:
            break
    return True

def scroll_detail_panel(page, max_rounds=10):
    panel = page.locator("div[role='main']")
    if panel.count() == 0:
        return
    last = 0
    for _ in range(max_rounds):
        try:
            panel.first.evaluate("(el)=>el.scrollBy(0, el.scrollHeight)")
            time.sleep(0.12)
            h = panel.first.evaluate("(el)=>el.scrollHeight")
            if h == last:
                break
            last = h
        except:
            break

# ======================
# BASIC FIELDS
# ======================

def get_address(page):
    for s in [
        "button[data-item-id='address']",
        "button[aria-label^='Địa chỉ']",
        "div[aria-label^='Địa chỉ']",
    ]:
        loc = page.locator(s)
        if loc.count() > 0:
            return (loc.first.text_content() or "").strip()
    return ""

def get_phone(page):
    loc = page.locator("button[data-item-id^='phone']")
    return (loc.first.text_content() or "").strip() if loc.count() > 0 else ""

def get_website(page):
    loc = page.locator("a[data-item-id='authority']")
    return loc.first.get_attribute("href") if loc.count() > 0 else ""

def get_located_in(page):
    try:
        loc = page.locator("text=/Nằm ở:/")
        if loc.count() > 0:
            txt = loc.first.text_content() or ""
            return txt.replace("Nằm ở:", "").strip()
    except:
        pass
    return ""

# ======================
# OPENING HOURS
# ======================

def expand_opening_hours_block(page):
    scroll_detail_panel(page)
    candidates = [
        "button:has-text('Giờ')",
        "button:has-text('Mở cửa')",
        "button:has-text('Đóng cửa')",
        "div[role='button']:has-text('Giờ')",
        "div[role='button']:has-text('Mở cửa')",
        "div[role='button']:has-text('Đóng cửa')",
        "button[data-item-id='oh']",
        "div[data-item-id='oh']",
    ]
    for s in candidates:
        loc = page.locator(s)
        if loc.count() > 0:
            try:
                loc.first.scroll_into_view_if_needed()
                time.sleep(0.2)
                loc.first.click(force=True)
                page.wait_for_selector("text=/Thứ|Chủ/", timeout=6000)
                time.sleep(0.2)
                return True
            except:
                pass
    return False

def scan_opening_hours_by_rows(page):
    hours = {}
    try:
        panel = page.locator("div[role='main']")
        if panel.count() > 0:
            panel_text = (panel.first.text_content() or "")
            pattern = r"(Thứ\s+(?:Hai|Ba|Tư|Năm|Sáu|Bảy)|Chủ\s+Nhật)\s*([0-9:–-]+\s*(?:–|-)\s*[0-9:–-]+|Đóng cửa|Mở 24 giờ)"
            matches = re.findall(pattern, panel_text, re.IGNORECASE)
            for d, t in matches:
                if t:
                    hours[d.strip()] = t.strip()
            if hours:
                return hours
    except:
        pass
    return hours

def get_opening_hours_today(page):
    for s in [
        "button[data-item-id='oh']",
        "div[data-item-id='oh']",
        "button[aria-label*='Mở cửa']",
        "button[aria-label*='Đóng cửa']",
    ]:
        loc = page.locator(s)
        if loc.count() > 0:
            return (loc.first.text_content() or "").strip()
    return ""

# ======================
# LIGHT FILTER
# ======================

def get_basic_poi_for_filter(page):
    try:
        name = (page.locator("h1").first.text_content() or "").strip()
        lat, lng = extract_latlng_from_url(page.url)
        return name, lat, lng
    except:
        return None, "", ""

# ======================
# PARSE POI FULL
# ======================

def parse_current_poi(page, keyword):
    # Lấy địa chỉ từ POI
    address = get_address(page)

    # Nếu địa chỉ không có, in ra thông báo
    if not address:
        print(f"🚫 Không có địa chỉ, tìm kiếm theo keyword: {keyword}")
    else:
        print(f"📍 Địa chỉ: {address}. Tiến hành phân vùng tìm kiếm.")

    try:
        name = (page.locator("h1").first.text_content() or "").strip()
        if not name:
            return None
    except:
        return None

    scroll_detail_panel(page)
    located_in = get_located_in(page)
    phone = normalize_phone_vn(get_phone(page))
    website = get_website(page)
    opening_today = get_opening_hours_today(page)

    opening_full = {}
    if expand_opening_hours_block(page):
        opening_full = scan_opening_hours_by_rows(page)

    # Chuẩn hóa giờ mở cửa theo thứ tự từ Thứ Hai đến Chủ Nhật
    opening_full = normalize_opening_hours(opening_full)

    lat, lng = extract_latlng_from_url(page.url)

    return {
        "keyword": keyword,
        "name": name,
        "address": address,
        "located_in": located_in,
        "phone": phone,
        "website": website,
        "opening_hours_today": opening_today,
        "opening_hours_full": json.dumps(opening_full, ensure_ascii=False),
        "lat": lat,
        "lng": lng,
        "url": page.url
    }


def crawl_google_maps_keyword(page, keyword, center_lat=None, center_lng=None, radius_km=0.5):
    results = []
    sb = page.wait_for_selector("input[role='combobox']", timeout=15000)
    sb.click()
    sb.fill(keyword)
    sb.press("Enter")
    time.sleep(3)

    has_feed = scroll_results(page)
    if has_feed:
        links = get_place_links_from_list(page)
    else:
        # Không có list → có thể đang ở trang chi tiết luôn
        links = [page.url]

    for i, l in enumerate(links, 1):
        try:
            page.goto(l, timeout=60000)
            time.sleep(0.8)

            name, lat, lng = get_basic_poi_for_filter(page)
            if not name:
                continue

            if center_lat and center_lng and lat and lng:
                d = distance_km(center_lat, center_lng, float(lat), float(lng))
                if d > radius_km:
                    print(f"⛔ Ngoài {radius_km}km:", name, round(d * 1000), "m")
                    continue  # không parse full

            poi = parse_current_poi(page, keyword)
            if poi:
                results.append(poi)
                print(f"✔ {keyword} | {i}: {poi['name']} → DỪNG keyword")
                break  # dừng keyword khi gặp POI hợp lệ

        except Exception as e:
            print("❌", keyword, e)

    return results

# ======================
# GET LINKS
# ======================

def get_place_links_from_list(page):
    try:
        return page.eval_on_selector_all(
            "a[href*='/maps/place/']",
            "els => [...new Set(els.map(e => e.href))]"
        )
    except:
        return []

# ======================
# FOCUS LOCATION
# ======================

def normalize_opening_hours(opening_hours):
    # Các ngày trong tuần theo thứ tự chuẩn
    week_days = ["Thứ Hai", "Thứ Ba", "Thứ Tư", "Thứ Năm", "Thứ Sáu", "Thứ Bảy", "Chủ Nhật"]

    # Tạo dictionary với giá trị mặc định là ""
    normalized_hours = {day: "" for day in week_days}

    # Lấp đầy dữ liệu cào vào dictionary
    for day, hours in opening_hours.items():
        if day in normalized_hours:
            normalized_hours[day] = hours

    return normalized_hours


# Test với dữ liệu bạn cào
opening_hours_full = {
    "Thứ Năm": "08:00–22:00",
    "Thứ Sáu": "08:00–22:00",
    "Thứ Bảy": "08:00–22:00",
    "Chủ Nhật": "08:00–22:00",
    "Thứ Hai": "08:00–22:00",
    "Thứ Ba": "08:00–22:00",
    "Thứ Tư": "08:00–22:00"
}

normalized_opening_hours = normalize_opening_hours(opening_hours_full)
print(normalized_opening_hours)


def focus_location_with_radius(page, loc_text, radius_km=0.5):
    try:
        # Thay vì lat, long, giờ ta sử dụng địa chỉ text để tìm kiếm
        page.goto(f"https://www.google.com/maps/search/{loc_text}", timeout=60000)
        time.sleep(2)
        # Sau khi tìm kiếm địa chỉ trên Maps, lấy lại tọa độ từ kết quả tìm kiếm
        lat_lng = page.locator("meta[name='place:location']").get_attribute("content")
        if lat_lng:
            lat, lng = lat_lng.split(",")
            lat = float(lat)
            lng = float(lng)
            return lat, lng
    except:
        return None, None

# ======================
# AUTOSAVE XLSX
# ======================

FIELDS = [
    "keyword", "name", "address", "located_in", "phone", "website",
    "opening_hours_today", "opening_hours_full", "lat", "lng", "url"
]

def save_xlsx(path, data):
    wb = Workbook()
    ws = wb.active
    ws.append(FIELDS)
    for row in data:
        ws.append([row.get(k, "") for k in FIELDS])
    wb.save(path)

# ======================
# CRAWL 1 KEYWORD
# ======================

def crawl_google_maps_keyword(page, keyword, center_lat=None, center_lng=None, radius_km=0.5):
    results = []
    sb = page.wait_for_selector("input[role='combobox']", timeout=15000)
    sb.click()
    sb.fill(keyword)
    sb.press("Enter")
    time.sleep(3)

    has_feed = scroll_results(page)
    if has_feed:
        links = get_place_links_from_list(page)
    else:
        # Không có list → có thể đang ở trang chi tiết luôn
        links = [page.url]

    for i, l in enumerate(links, 1):
        try:
            page.goto(l, timeout=60000)
            time.sleep(0.8)

            name, lat, lng = get_basic_poi_for_filter(page)
            if not name:
                continue

            if center_lat and center_lng and lat and lng:
                d = distance_km(center_lat, center_lng, float(lat), float(lng))
                if d > radius_km:
                    print(f"⛔ Ngoài {radius_km}km:", name, round(d*1000), "m")
                    continue   # không parse full

            poi = parse_current_poi(page, keyword)
            if poi:
                results.append(poi)
                print(f"✔ {keyword} | {i}: {poi['name']} → DỪNG keyword")
                break   # dừng keyword khi gặp POI hợp lệ

        except Exception as e:
            print("❌", keyword, e)
    return results

# ======================
# MAIN
# ======================

if __name__ == "__main__":
    script_dir = os.path.dirname(os.path.abspath(__file__))

    root = tk.Tk()
    root.withdraw()
    path = filedialog.askopenfilename(title="Chọn file Excel", filetypes=[("Excel", "*.xlsx *.xls")])
    if not path:
        print("❌ Chưa chọn file")
        exit()

    print("📄 File Excel:", path)
    wb = load_workbook(path)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    for i, h in enumerate(headers, 1):
        print(i, h)

    key_idx = int(input("🔎 Cột keyword: ")) - 1
    loc_idx = int(input("📍 Cột địa chỉ (dạng text): ")) - 1  # Chọn cột chứa địa chỉ

    rows_data = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        kw = clean_cell(r[key_idx]) if key_idx < len(r) else ""
        loc = clean_cell(r[loc_idx]) if loc_idx < len(r) else ""  # Cột chứa địa chỉ
        if kw:
            rows_data.append((kw, loc))

    radius = 0.5
    autosave_path = os.path.join(script_dir, "autosave_temp.xlsx")
    all_results = []

    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            user_data_dir="google_profile",
            executable_path=r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            headless=False,
            locale="vi-VN",
            viewport={"width": 1280, "height": 800}
        )
        page = context.new_page()
        page.goto("https://www.google.com/maps?hl=vi", timeout=60000)
        time.sleep(2)

        for i, (k, loc) in enumerate(rows_data, 1):
            print(f"\n▶️ {i}/{len(rows_data)}: {k}")
            center_lat = center_lng = None

            if loc:
                print("📍 Keyword chung → dùng location:", loc)
                center_lat, center_lng = focus_location_with_radius(page, loc, radius)
            else:
                print("🎯 Keyword chi tiết → không cần location")

            results = crawl_google_maps_keyword(page, k, center_lat, center_lng, radius)
            all_results.extend(results)
            if all_results:
                save_xlsx(autosave_path, all_results)
                print("💾 Autosave:", autosave_path)
            time.sleep(1)

        context.close()

    out = os.path.join(script_dir, f"googlemaps_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    save_xlsx(out, all_results)
    print("✅ File cuối:", out)
    print("🧾 File autosave:", autosave_path)
