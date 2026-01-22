import time, csv, os, re, json
from datetime import datetime
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

# ======================
# CLEAN CELL
# ======================
def clean_cell(v):
    return "" if v is None else str(v).strip()

# ======================
# NORMALIZE PHONE (VN)
# ======================
def normalize_phone_vn(phone):
    if not phone: return ""
    d = re.sub(r"\D","",phone)
    if not d: return ""
    if d.startswith("84"): return f"+{d}"
    if d.startswith("0"): return f"+84{d[1:]}"
    return f"+{d}"

# ======================
# EXTRACT lat,lng from URL
# ======================
def extract_latlng_from_url(url):
    m = re.search(r"@(-?\d+\.\d+),(-?\d+\.\d+)", url)
    return (m.group(1), m.group(2)) if m else ("","")

# ======================
# SCROLL RESULTS LIST
# ======================
def scroll_results(page, max_rounds=40):
    feed = page.locator("div[role='feed']")
    if feed.count() == 0: return False
    last = 0
    for _ in range(max_rounds):
        try:
            feed.first.evaluate("(el)=>el.scrollBy(0, el.scrollHeight)")
            time.sleep(0.5)
            h = feed.first.evaluate("(el)=>el.scrollHeight")
            if h == last: break
            last = h
        except:
            break
    return True

# ======================
# SCROLL DETAIL PANEL
# ======================
def scroll_detail_panel(page, max_rounds=15):
    panel = page.locator("div[role='main']")
    if panel.count() == 0: return
    last = 0
    for _ in range(max_rounds):
        try:
            panel.first.evaluate("(el)=>el.scrollBy(0, el.scrollHeight)")
            time.sleep(0.2)
            h = panel.first.evaluate("(el)=>el.scrollHeight")
            if h == last: break
            last = h
        except:
            break

# ======================
# BASIC FIELDS
# ======================
def get_address(page):
    for s in [
        "button[data-item-id='address']",
        "button[aria-label^='Địa chỉ']",
        "div[aria-label^='Địa chỉ']",
    ]:
        loc = page.locator(s)
        if loc.count() > 0:
            return (loc.first.text_content() or "").strip()
    return ""

def get_phone(page):
    loc = page.locator("button[data-item-id^='phone']")
    return (loc.first.text_content() or "").strip() if loc.count()>0 else ""

def get_website(page):
    loc = page.locator("a[data-item-id='authority']")
    return loc.first.get_attribute("href") if loc.count()>0 else ""

# ======================
# OPENING HOURS (EXPAND + COPY DETECTION)
# ======================
def expand_opening_hours_block(page):
    scroll_detail_panel(page)
    candidates = [
        "button:has-text('Giờ')",
        "button:has-text('Mở cửa')",
        "button:has-text('Đóng cửa')",
        "div[role='button']:has-text('Giờ')",
        "div[role='button']:has-text('Mở cửa')",
        "div[role='button']:has-text('Đóng cửa')",
        "button[data-item-id='oh']",
        "div[data-item-id='oh']",
    ]
    for s in candidates:
        loc = page.locator(s)
        if loc.count() > 0:
            try:
                loc.first.scroll_into_view_if_needed()
                time.sleep(0.3)
                loc.first.click(force=True)
                page.wait_for_selector("text=/Thứ|Chủ/", timeout=8000)
                time.sleep(0.3)
                return True
            except:
                pass
    return False

def read_clipboard(page):
    try:
        # Dùng evaluate_handle để xử lý async code đúng cách
        result = page.evaluate("""async () => {
            try {
                return await navigator.clipboard.readText();
            } catch (e) {
                console.log('Clipboard error:', e);
                return '';
            }
        }""")
        return result if result else ""
    except Exception as e:
        print(f"⚠️ Lỗi đọc clipboard: {e}")
        return ""

def scan_opening_hours_by_rows(page):
    hours = {}
    
    print(f"\n🔎 Scan opening hours:")
    
    # Cách 1: Thử lấy từ panel chính
    try:
        panel = page.locator("div[role='main']")
        if panel.count() > 0:
            panel_text = (panel.first.text_content() or "").strip()
            print(f"   Panel text length: {len(panel_text)}")
            
            # Parse toàn bộ bằng regex - tìm tất cả dòng ngày + giờ
            # Lưu ý: Không có khoảng trắng giữa ngày và giờ (Thứ Ba10:00–22:00)
            pattern = r"(Thứ\s+(?:Hai|Ba|Tư|Năm|Sáu|Bảy)|Chủ\s+Nhật)\s*([0-9:–-]+\s*(?:–|-)\s*[0-9:–-]+|Đóng cửa|Mở 24 giờ)"
            matches = re.findall(pattern, panel_text, re.IGNORECASE)
            
            print(f"   Tìm được {len(matches)} match từ panel")
            for day, time_info in matches:
                day = day.strip()
                time_info = time_info.strip()
                if time_info:
                    print(f"      ✅ {day} | {time_info}")
                    hours[day] = time_info
            
            if hours:
                print(f"\n📊 Kết quả từ panel: {hours}\n")
                return hours
    except Exception as e:
        print(f"   ⚠️ Lỗi lấy từ panel: {e}")
    
    # Cách 2: Nếu panel không được, tìm từng row
    print(f"\n   Thử cách 2: Parse từ container row...")
    rows = page.locator("text=/Thứ|Chủ/")
    
    print(f"   Tìm thấy {rows.count()} dòng giờ")
    
    if rows.count() == 0:
        print(f"   ⚠️ Không tìm thấy dòng nào, thử scroll lại...")
        scroll_detail_panel(page)
        rows = page.locator("text=/Thứ|Chủ/")
        print(f"   Sau scroll: {rows.count()} dòng")

    for i in range(rows.count()):
        try:
            row = rows.nth(i)
            row.scroll_into_view_if_needed()
            time.sleep(0.2)

            # Tìm container cha chứa dòng này
            try:
                row_container = row.evaluate("""(el) => {
                    let parent = el.parentElement;
                    // Tìm parent chứa cả ngày và giờ
                    while (parent && parent.textContent.length < 200) {
                        parent = parent.parentElement;
                    }
                    return parent ? parent.textContent : el.textContent;
                }""").strip()
            except:
                row_container = row.text_content() or ""
            
            row_container = re.sub(r"\s+", " ", row_container)
            
            # Parse từ container - không yêu cầu khoảng trắng sau ngày
            pattern = r"(Thứ\s+(?:Hai|Ba|Tư|Năm|Sáu|Bảy)|Chủ\s+Nhật)\s*([0-9:–-]+\s*(?:–|-)\s*[0-9:–-]+|Đóng cửa|Mở 24 giờ)"
            m = re.search(pattern, row_container, re.IGNORECASE)
            
            if m:
                day = m.group(1).strip()
                time_range = m.group(2).strip()
                
                if time_range and len(time_range) > 1:
                    print(f"   [{i+1}] ✅ {day} | {time_range}")
                    hours[day] = time_range

            # Click copy button nếu có
            copy_btn = row.locator("button:has-text('Sao chép'), button:has-text('Copy')")
            if copy_btn.count() > 0:
                try:
                    copy_btn.first.click(force=True)
                    time.sleep(0.2)
                except:
                    pass
            
        except Exception as e:
            print(f"   ❌ Lỗi row {i}: {e}")
    
    print(f"\n📊 Kết quả giờ mở cửa: {hours}\n")
    return hours

def get_opening_hours_today(page):
    for s in [
        "button[data-item-id='oh']",
        "div[data-item-id='oh']",
        "button[aria-label*='Mở cửa']",
        "button[aria-label*='Đóng cửa']",
    ]:
        loc = page.locator(s)
        if loc.count() > 0:
            return (loc.first.text_content() or "").strip()
    return ""

# ======================
# PARSE POI DETAIL
# ======================
def parse_current_poi(page, keyword):
    try:
        name = (page.locator("h1").first.text_content() or "").strip()
        if not name: return None
    except:
        return None

    scroll_detail_panel(page)

    address = get_address(page)
    phone = normalize_phone_vn(get_phone(page))
    website = get_website(page)
    opening_today = get_opening_hours_today(page)

    opening_full = {}
    opened = expand_opening_hours_block(page)
    print("DEBUG opened hours block:", opened)
    if opened:
        opening_full = scan_opening_hours_by_rows(page)

    lat, lng = extract_latlng_from_url(page.url)

    return {
        "keyword": keyword,
        "name": name,
        "address": address,
        "phone": phone,
        "website": website,
        "opening_hours_today": opening_today,
        "opening_hours_full": json.dumps(opening_full, ensure_ascii=False),
        "lat": lat,
        "lng": lng,
        "url": page.url
    }

# ======================
# GET LINKS FROM LIST
# ======================
def get_place_links_from_list(page):
    try:
        return page.eval_on_selector_all(
            "a[href*='/maps/place/']",
            "els=>[...new Set(els.map(e=>e.href))]"
        )
    except:
        return []

# ======================
# CRAWL 1 KEYWORD
# ======================
def crawl_google_maps_keyword(page, keyword):
    results = []
    sb = page.wait_for_selector("input[role='combobox']", timeout=15000)
    sb.click(); sb.fill(keyword); sb.press("Enter")
    time.sleep(3.5)

    has_feed = scroll_results(page)
    links = get_place_links_from_list(page) if has_feed else []

    if not links:
        poi = parse_current_poi(page, keyword)
        if poi: results.append(poi)
        return results

    for i, l in enumerate(links, 1):
        try:
            page.goto(l, timeout=60000)
            time.sleep(1.5)
            poi = parse_current_poi(page, keyword)
            if poi:
                results.append(poi)
                print(f"✔ {keyword} | {i}: {poi['name']}")
        except Exception as e:
            print("❌", keyword, e)
    return results

# ======================
# SAVE CSV
# ======================
def save_all_to_csv(out, data):
    print(f"\n🔍 Chuẩn bị save CSV...")
    print(f"📊 Tổng records: {len(data) if data else 0}")
    
    if not data:
        print(f"⚠️ Không có dữ liệu, tạo CSV trống")
        data = [{
            "keyword": "",
            "name": "",
            "address": "",
            "phone": "",
            "website": "",
            "opening_hours_today": "",
            "opening_hours_full": "",
            "lat": "",
            "lng": "",
            "url": ""
        }]
    
    out_path = os.path.abspath(out)
    print(f"📁 Lưu tại: {out_path}")
    
    try:
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=data[0].keys())
            w.writeheader()
            w.writerows(data)
        print(f"✅ Lưu CSV thành công!")
        print(f"📊 Số dòng dữ liệu: {len(data)}")
    except Exception as e:
        print(f"❌ Lỗi lưu CSV: {e}")

# ======================
# MAIN (EXCEL: each row = keyword)
# ======================
if __name__ == "__main__":
    folder = input("📂 Folder Excel: ").strip()
    file = input("📄 File Excel: ").strip()
    path = os.path.join(folder, file)

    wb = load_workbook(path); ws = wb.active
    headers = [c.value for c in ws[1]]
    for i, h in enumerate(headers, 1): print(i, h)
    key_idx = int(input("🔎 STT cột keyword: ")) - 1

    keywords = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if key_idx < len(r):
            k = clean_cell(r[key_idx])
            if k: keywords.append(k)

    all_results = []
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, slow_mo=100)
        ctx = browser.new_context(
            locale="vi-VN",
            viewport={"width":1280,"height":800},
            permissions=["clipboard-read","clipboard-write"]
        )
        page = ctx.new_page()
        page.goto("https://www.google.com/maps?hl=vi", timeout=60000)
        time.sleep(2)

        for i, k in enumerate(keywords, 1):
            print(f"\n▶️ {i}/{len(keywords)}: {k}")
            results = crawl_google_maps_keyword(page, k)
            print(f"  └─ Lấy được {len(results)} kết quả")
            all_results.extend(results)
            time.sleep(1)
        browser.close()

    print(f"\n📊 TỔNG KẾT:")
    print(f"   Tổng keywords: {len(keywords)}")
    print(f"   Tổng POI lấy được: {len(all_results)}")
    
    out = f"googlemaps_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    save_all_to_csv(out, all_results)
    
    if os.path.exists(out):
        file_size = os.path.getsize(out)
        print(f"✅ File saved: {out}")
        print(f"   Kích thước: {file_size} bytes")
    else:
        print(f"❌ File không tìm thấy: {out}")
