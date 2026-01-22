import json
import os
import time
import requests
from openpyxl import Workbook, load_workbook

# ===================== MAP4D CONFIG =====================
MAP4D_API_KEY = "93d393d0f6507ee00b62fe01db7430fa"
MAP4D_API_URL = "https://api.map4d.vn/sdk/v2/geocode"

def get_admin_levels_from_map4d(lat, lon):
    """Lấy Province / District / Ward từ Map4D"""
    try:
        url = f"{MAP4D_API_URL}?location={lat}%2C{lon}&key={MAP4D_API_KEY}"
        res = requests.get(url, timeout=10)

        if res.status_code != 200:
            print(f"⚠️ API lỗi {res.status_code} ({lat}, {lon})")
            return None, None, None

        data = res.json()
        results = data.get("result", [])
        if not results:
            return None, None, None

        comp = results[0].get("addressComponents", [])
        old_comp = results[0].get("oldAddressComponents", [])

        def extract(components):
            p = d = w = None
            for c in components:
                types = c.get("types", [])
                if "admin_level_2" in types:
                    p = c.get("name")
                elif "admin_level_3" in types:
                    d = c.get("name")
                elif "admin_level_4" in types:
                    w = c.get("name")
            return p, d, w

        province, district, ward = extract(comp)

        if not (province and district and ward):
            p2, d2, w2 = extract(old_comp)
            province = province or p2
            district = district or d2
            ward = ward or w2

        return province, district, ward

    except Exception as e:
        print(f"⚠️ Lỗi Map4D ({lat}, {lon}): {e}")
        return None, None, None


def filter_poi_with_dedup_and_excel():
    # Bước nhập thư mục và file chỉ thực hiện 1 lần ban đầu
    folder_path = input("📂 Nhập đường dẫn thư mục chứa GeoJSON: ").strip()
    if not os.path.isdir(folder_path):
        print("❌ Thư mục không tồn tại")
        return

    input_filename = input("📄 Nhập tên file GeoJSON: ").strip()
    input_path = os.path.join(folder_path, input_filename)
    if not os.path.isfile(input_path):
        print("❌ File không tồn tại")
        return

    output_folder = input("📁 Nhập thư mục đầu ra để lưu file kết quả: ").strip()
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)  # Tạo thư mục nếu chưa tồn tại

    while True:
        # Bước nhập keyword
        raw_keywords = input(
            "🔎 Nhập keyword (cách nhau bằng dấu phẩy): "
        ).strip()
        if not raw_keywords:
            print("❌ Chưa nhập keyword")
            continue  # Nếu không nhập keyword, quay lại bước này

        keywords = [k.strip().upper() for k in raw_keywords.split(",") if k.strip()]

        # Bước xác định cách lọc trùng
        print("\n⚙️ Chọn cách lọc trùng:")
        print("1️⃣  Theo name")
        print("2️⃣  Theo tọa độ")
        print("3️⃣  Theo name + tọa độ")
        choice = input("👉 Chọn (1/2/3): ").strip()

        # Bước xử lý dữ liệu
        with open(input_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        features = data.get("features", [])

        # Lọc theo keyword
        filtered_features = [
            ft for ft in features
            if any(
                kw in ft.get("properties", {}).get("name", "").upper()
                for kw in keywords
            )
        ]

        print(f"\n🔍 Tìm được {len(filtered_features)} POI")

        # Xử lý loại trùng
        unique = []
        seen = set()

        for ft in filtered_features:
            props = ft.get("properties", {})
            geom = ft.get("geometry", {})

            name = props.get("name", "").strip()
            coords = tuple(geom.get("coordinates", []))

            if choice == "1":
                key = name
            elif choice == "2":
                key = coords
            else:
                key = (name, coords)

            if key not in seen:
                seen.add(key)
                unique.append(ft)

        print(f"✅ Sau khi loại trùng: {len(unique)} POI")

        # Lưu tên file theo từ khóa
        safe_keyword = keywords[0].replace(" ", "_")

        geojson_out = os.path.join(output_folder, f"poi_{safe_keyword}_dedup.geojson")
        excel_out = os.path.join(output_folder, f"poi_{safe_keyword}_dedup.xlsx")

        # Xuất file GeoJSON
        with open(geojson_out, "w", encoding="utf-8") as f:
            json.dump(
                {"type": "FeatureCollection", "features": unique},
                f,
                ensure_ascii=False,
                indent=2
            )

        print(f"📁 Đã xuất GeoJSON: {geojson_out}")

        # Bước xuất Excel
        if input("\n📊 Xuất Excel? (Y/N): ").strip().upper() != "Y":
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "POI"

        ws.append([
            "STT", "Name", "Class", "Layer",
            "Latitude", "Longitude",
            "Province", "District", "Ward"
        ])

        for idx, ft in enumerate(unique, 1):
            props = ft.get("properties", {})
            coords = ft.get("geometry", {}).get("coordinates", [None, None])

            lon, lat = coords[0], coords[1]

            ws.append([
                idx,
                props.get("name", ""),
                props.get("class", ""),
                props.get("vt_layer", ""),
                lat,
                lon,
                "", "", ""
            ])

        wb.save(excel_out)
        print(f"✅ Đã xuất Excel: {excel_out}")

        # Bước Tra Map4D
        map4d_choice = input("\n🌍 Tra Map4D (tỉnh/huyện/xã)? (Y/N): ").strip().upper()
        if map4d_choice == "N":
            # Nếu trả lời "No", tiếp tục hỏi có muốn tiếp tục không
            continue_choice = input("\n🔁 Bạn có muốn tiếp tục không? (Yes để quay lại bước 4, No để kết thúc): ").strip().lower()
            if continue_choice != 'yes':
                print("🚀 Kết thúc chương trình.")
                break  # Kết thúc chương trình nếu người dùng chọn 'No'
            else:
                continue  # Quay lại bước 4 nếu người dùng chọn 'Yes'

        if map4d_choice == "Y":
            wb = load_workbook(excel_out)
            ws = wb.active

            for row in range(2, ws.max_row + 1):
                lat = ws.cell(row=row, column=5).value
                lon = ws.cell(row=row, column=6).value

                if not (lat and lon):
                    continue

                province, district, ward = get_admin_levels_from_map4d(lat, lon)

                ws.cell(row=row, column=7).value = province or "Không xác định"
                ws.cell(row=row, column=8).value = district or ""
                ws.cell(row=row, column=9).value = ward or ""

                print(f"📍 {lat},{lon} → {province} | {district} | {ward}")

                time.sleep(0.2)  # tránh rate limit

            wb.save(excel_out)
            print("\n🎉 HOÀN TẤT TOÀN BỘ QUY TRÌNH")
            break  # Kết thúc chương trình sau khi tra cứu Map4D


if __name__ == "__main__":
    filter_poi_with_dedup_and_excel()
